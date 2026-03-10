import pandas as pd
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, Border, Alignment
import threading
import time


# ----------------- helpers -----------------

def try_parse_input_date(s):
    s = s.strip()
    formats = [
        "%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d",
        "%m/%d/%Y", "%d/%m/%Y",
        "%m-%d-%Y", "%d-%m-%Y",
        "%Y.%d.%m"
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # fallback to pandas fuzzy parse
    try:
        d = pd.to_datetime(s, dayfirst=False, errors="coerce")
        if pd.isna(d):
            return None
        return d.date()
    except Exception:
        return None


def excel_col_to_index(letter):
    return column_index_from_string(letter) - 1


def ensure_openpyxl_installed():
    try:
        import openpyxl  # noqa: F401
    except Exception:
        raise RuntimeError("Package 'openpyxl' is required. Install with: pip install openpyxl")


# ----------------- core processing -----------------

def process_files(db_path, shipment_path, date_inputs, output_dir, progress_callback=None):
    """
    Core file processing function (optimized version)
    progress_callback: Optional callback function for progress updates
    """

    def log(msg):
        print(msg)  # Console output
        if progress_callback:
            progress_callback(msg)

    start_time = time.time()
    ensure_openpyxl_installed()

    log("Reading database file...")

    # read database sheet - only read rows with data
    try:
        log("Opening database file...")

        # 直接读取前10000行（通常实际数据不会超过这个数）
        # 这样可以避免扫描100万行的问题
        max_rows_to_read = 10000  # 可以根据实际情况调整这个数字

        log(f"Reading first {max_rows_to_read:,} rows from '2025 Orders' sheet...")
        df_db = pd.read_excel(db_path, sheet_name="2025 Orders", engine="openpyxl",
                              dtype=str, nrows=max_rows_to_read)

        log(f"✓ Initial read completed, got {len(df_db):,} rows")

    except Exception as e:
        raise RuntimeError(f"Failed to read database file: {e}")

    df_db = df_db.fillna("")

    # Filter out completely blank rows
    df_db = df_db[df_db.astype(str).apply(lambda x: x.str.strip()).ne("").any(axis=1)]

    log(f"✓ Loaded {len(df_db):,} valid data rows (took {time.time() - start_time:.1f}s)")

    if df_db.shape[1] < 16:
        raise RuntimeError(
            f"Thermo Database '2025 Orders' must have at least 16 columns, currently has {df_db.shape[1]} columns.")

    # determine M column label (prefer header 'EST. DELIVERY DATE')
    date_col_label = None
    for c in df_db.columns:
        if str(c).strip().upper() == "EST. DELIVERY DATE".upper():
            date_col_label = c
            break
    if date_col_label is None:
        for c in df_db.columns:
            if str(c).strip().upper() == "M":
                date_col_label = c
                break
    if date_col_label is None:
        date_col_label = df_db.columns[12]

    # check for blank cells in M column and filter them out
    log("Checking for blank dates...")
    blank_mask = df_db[date_col_label].astype(str).str.strip() == ""
    if blank_mask.any():
        blank_count = blank_mask.sum()
        first_few = df_db[blank_mask].head(5).index + 2
        rows = ", ".join(str(i) for i in first_few)
        log(f"⚠ Warning: Found {blank_count} rows with blank dates (Excel rows: {rows}, etc.), skipping these rows")
        # Filter out rows with blank dates
        df_db = df_db[~blank_mask].copy()
        log(f"✓ After filtering: {len(df_db):,} valid data rows remaining")

    log(f"Parsing {len(df_db):,} rows of date data (this may take 1-2 minutes)...")
    parse_start = time.time()

    # parse M column to date - optimized date parsing
    try:
        parsed = pd.to_datetime(df_db[date_col_label], errors="coerce", format='mixed')
    except Exception as e:
        raise RuntimeError(f"Failed to parse dates in column '{date_col_label}': {e}")

    log(f"✓ Date parsing completed (took {time.time() - parse_start:.1f}s)")

    if parsed.isna().any():
        bad_count = parsed.isna().sum()
        bad_idx = parsed[parsed.isna()].head(5).index
        rows = ", ".join(str(i + 2) for i in bad_idx)
        log(f"⚠ Warning: Found {bad_count} rows with invalid dates (Excel rows: {rows}, etc.), skipping these rows")
        # Filter out rows with invalid dates
        df_db = df_db[~parsed.isna()].copy()
        parsed = parsed[~parsed.isna()]
        log(f"✓ After filtering: {len(df_db):,} valid data rows remaining")

    df_db["_PARSED_DATE_"] = parsed.dt.date

    # parse user input dates
    log("Parsing input dates...")
    parsed_dates = []
    parse_errors = []
    for s in date_inputs:
        d = try_parse_input_date(s)
        if d is None:
            parse_errors.append(s)
        else:
            parsed_dates.append((s, d))
    if parse_errors:
        raise RuntimeError("Unable to parse input date(s): " + ", ".join(parse_errors) +
                           ". Example formats: 2025.12.12 or 2025-12-12 or 12/12/2025")

    if not parsed_dates:
        raise RuntimeError("No valid dates provided.")

    log(f"Searching for matching records for {len(parsed_dates)} date(s)...")
    search_start = time.time()

    # Use isin for batch matching (much faster than individual matching)
    target_dates = [d for _, d in parsed_dates]
    matched_df = df_db[df_db["_PARSED_DATE_"].isin(target_dates)].copy()

    log(f"✓ Search completed, found {len(matched_df):,} matching records (took {time.time() - search_start:.1f}s)")

    if matched_df.empty:
        raise RuntimeError("No matching records found for any of the provided dates. Dates checked: " +
                           ", ".join([p[0] for p in parsed_dates]))

    # Count matches per date
    no_match_inputs = []
    for raw_input, target_date in parsed_dates:
        count = (matched_df["_PARSED_DATE_"] == target_date).sum()
        if count == 0:
            no_match_inputs.append(raw_input)
        else:
            log(f"  Date {raw_input}: {count} records")

    # open shipment workbook
    log("Opening shipment file...")
    shipment_ext = os.path.splitext(shipment_path)[1].lower()
    keep_vba = shipment_ext == ".xlsm"
    try:
        wb = load_workbook(shipment_path, keep_vba=keep_vba)
    except PermissionError:
        raise RuntimeError("Cannot open shipment file. The file may be open in Excel. Please close it and try again.")
    except Exception as e:
        raise RuntimeError(f"Failed to open shipment template: {e}")

    if "Invoice" not in wb.sheetnames:
        raise RuntimeError("Shipment workbook does not contain a sheet named 'Invoice'.")

    ws = wb["Invoice"]
    max_col = ws.max_column

    # mapping to Excel columns
    mapping = {
        "D": excel_col_to_index("D"),
        "E": excel_col_to_index("E"),
        "F": excel_col_to_index("F"),
        "G": excel_col_to_index("G"),
        "H": excel_col_to_index("H"),
        "I": excel_col_to_index("I"),
        "J": excel_col_to_index("J"),
        "K": excel_col_to_index("K"),
        "L": excel_col_to_index("L"),
        "M": excel_col_to_index("M"),
        "N": excel_col_to_index("N"),
        "O": excel_col_to_index("O"),
    }

    # DB column indices
    COL_B = 1
    COL_D = 3
    COL_E = 4
    COL_F = 5
    COL_G = 6
    COL_H = 7
    # I列后面插入了新列，所以原J列之后的都+1
    COL_N = 14  # O列（索引14）→ 映射到输出F列（用于文件名分组）
    COL_O = 15  # P列（索引15）→ 映射到输出M列

    log(f"Preparing {len(matched_df):,} rows of data...")

    # Batch build data (faster than row by row)
    rows_to_insert = []
    for idx, (_, row) in enumerate(matched_df.iterrows()):
        rv = [""] * max_col

        def v(col_idx):
            try:
                return str(row.iloc[col_idx])
            except Exception:
                return ""

        rv[mapping["D"]] = "Packing Slip "
        rv[mapping["E"]] = v(COL_H)
        rv[mapping["F"]] = v(COL_N)
        rv[mapping["G"]] = v(COL_E)
        rv[mapping["H"]] = "Thermo Fisher Scientific Chemicals"
        rv[mapping["I"]] = v(COL_D)
        rv[mapping["J"]] = v(COL_F)
        rv[mapping["K"]] = "EA"
        g_val = v(COL_G)
        rv[mapping["L"]] = "RT" if g_val.strip() == "0" else g_val
        rv[mapping["M"]] = v(COL_O)
        rv[mapping["N"]] = v(COL_B)
        rv[mapping["O"]] = v(COL_D)

        rows_to_insert.append(rv)

        if (idx + 1) % 1000 == 0:
            log(f"  Prepared {idx + 1:,}/{len(matched_df):,} rows...")

    log(f"✓ Data preparation complete, total {len(rows_to_insert):,} rows")

    # clear old data
    log("Clearing old data...")
    last_row = ws.max_row
    if last_row >= 3:
        try:
            ws.delete_rows(3, last_row - 2)
        except PermissionError:
            raise RuntimeError(
                "Cannot modify shipment file; file may be open in Excel. Please close Excel and try again.")
        except Exception as e:
            raise RuntimeError(f"Failed to clear Invoice rows: {e}")

    log(f"Writing {len(rows_to_insert):,} rows of new data (this may take several minutes)...")
    write_start = time.time()

    # Write new data
    start_row = 3
    for r_idx, row_vals in enumerate(rows_to_insert):
        excel_row = start_row + r_idx
        for col_idx, value in enumerate(row_vals, start=1):
            ws.cell(row=excel_row, column=col_idx, value=value)

        # Report progress every 5000 rows
        if (r_idx + 1) % 5000 == 0:
            elapsed = time.time() - write_start
            rate = (r_idx + 1) / elapsed
            remaining = (len(rows_to_insert) - r_idx - 1) / rate if rate > 0 else 0
            log(f"  Written {r_idx + 1:,}/{len(rows_to_insert):,} rows ({(r_idx + 1) / len(rows_to_insert) * 100:.1f}%) - Estimated {remaining:.0f}s remaining")

    log(f"✓ Data write completed (took {time.time() - write_start:.1f}s)")
    log("Saving shipment file...")
    save_start = time.time()

    # save workbook
    try:
        wb.save(shipment_path)
    except PermissionError:
        raise RuntimeError(
            "Failed to save shipment file. The file may be open in Excel. Please close Excel and try again.")
    except Exception as e:
        raise RuntimeError(f"Failed to save updated shipment file: {e}")

    log(f"✓ File save completed (took {time.time() - save_start:.1f}s)")
    log("Re-reading Invoice sheet for grouping...")

    # re-load Invoice
    try:
        df_ship = pd.read_excel(shipment_path, sheet_name="Invoice", engine="openpyxl", dtype=str).fillna("")
    except Exception as e:
        raise RuntimeError(f"Failed to read updated Invoice sheet: {e}")

    if df_ship.shape[1] <= 5 or df_ship.shape[1] <= 13:
        raise RuntimeError("Updated Invoice sheet has fewer columns than expected for grouping by positions F/N.")

    col_F_name = df_ship.columns[5]
    col_N_name = df_ship.columns[13]

    log("Exporting files by group...")
    export_start = time.time()

    # group and export
    os.makedirs(output_dir, exist_ok=True)
    exported_files = []
    grouped = df_ship.groupby(by=col_F_name)

    total_groups = sum(1 for _ in grouped)
    group_count = 0

    for f_val, gF in grouped:
        f_key = str(f_val).strip()
        if f_key == "":
            continue
        for n_val, gFN in gF.groupby(by=col_N_name):
            n_key = str(n_val).strip()
            fname = f"{f_key} - PO {n_key}.xlsx"
            out_path = os.path.join(output_dir, fname)
            try:
                # Create workbook with sheet named "Invoice"
                with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                    # First write headers to row 1, data to row 3 (with gap at row 2)
                    gFN.to_excel(writer, sheet_name="Invoice", index=False, startrow=2)

                    # Now copy headers from row 3 to row 1 and clear row 3
                    workbook = writer.book
                    worksheet = workbook["Invoice"]

                    # Get the formatting from row 3 (original header row)
                    original_fonts = {}
                    original_borders = {}
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=3, column=col_idx)
                        original_fonts[col_idx] = cell.font.copy() if cell.font else None
                        original_borders[col_idx] = cell.border.copy() if cell.border else None

                    # Copy headers from row 3 to row 1, apply formatting, and skip "Unnamed" columns
                    for col_idx in range(1, worksheet.max_column + 1):
                        header_value = worksheet.cell(row=3, column=col_idx).value
                        cell_row1 = worksheet.cell(row=1, column=col_idx)

                        if header_value and not str(header_value).startswith("Unnamed"):
                            cell_row1.value = header_value
                            # Apply the original formatting to row 1
                            if original_fonts[col_idx]:
                                cell_row1.font = original_fonts[col_idx]
                            if original_borders[col_idx]:
                                cell_row1.border = original_borders[col_idx]

                        # Clear the header from row 3 and remove its formatting
                        cell_row3 = worksheet.cell(row=3, column=col_idx)
                        cell_row3.value = None
                        cell_row3.font = Font()
                        cell_row3.border = Border()
                        cell_row3.alignment = Alignment()

                    # Shift data up by one row (from row 4 onwards to row 3 onwards)
                    # and remove formatting from data rows
                    max_row = worksheet.max_row
                    for row_idx in range(4, max_row + 1):
                        for col_idx in range(1, worksheet.max_column + 1):
                            source_cell = worksheet.cell(row=row_idx, column=col_idx)
                            target_cell = worksheet.cell(row=row_idx - 1, column=col_idx)

                            target_cell.value = source_cell.value
                            # Remove bold/border/alignment formatting from data cells
                            target_cell.font = Font()
                            target_cell.border = Border()
                            target_cell.alignment = Alignment()

                    # Delete the last row (which is now duplicate)
                    worksheet.delete_rows(max_row)

                exported_files.append(out_path)
                group_count += 1
                if group_count % 5 == 0:
                    log(f"  Exported {group_count} files...")
            except Exception as e:
                raise RuntimeError(f"Failed to export file {out_path}: {e}")

    total_time = time.time() - start_time
    log(f"✓ All complete! Exported {len(exported_files)} files (total time {total_time:.1f}s = {total_time / 60:.1f} min)")

    return {"exported_files": exported_files, "no_match_inputs": no_match_inputs}


# ----------------- GUI -----------------

def on_run():
    db_file = entry_db.get().strip()
    ship_file = entry_ship.get().strip()
    dates_text = entry_dates.get().strip()
    out_dir = entry_out.get().strip()

    if not db_file or not ship_file or not dates_text or not out_dir:
        messagebox.showerror("Missing Input",
                             "Please select database file, shipment file, enter date(s), and choose output folder.")
        return

    date_list = [d.strip() for d in dates_text.split(",") if d.strip()]
    if not date_list:
        messagebox.showerror("Date Error", "Please enter at least one date (comma separated if multiple).")
        return

    # Clear log window
    log_text.delete(1.0, tk.END)

    # Disable button and update status
    btn_run.config(state="disabled", text="Processing...")
    progress_label.config(text="Starting process...", fg="blue")
    root.update()

    # Execute processing in background thread
    def process_in_thread():
        try:
            def update_progress(msg):
                # Use after to update GUI in main thread
                root.after(0, lambda m=msg: update_log_and_progress(m))

            result = process_files(db_file, ship_file, date_list, out_dir, update_progress)
            # Show results in main thread after completion
            root.after(0, lambda: show_success(result))
        except Exception as e:
            # Show error in main thread after error
            root.after(0, lambda: show_error(str(e)))

    thread = threading.Thread(target=process_in_thread, daemon=True)
    thread.start()


def update_log_and_progress(msg):
    """Update both progress label and log window"""
    # Update progress label
    progress_label.config(text=msg)

    # Add to log window with timestamp
    timestamp = datetime.now().strftime("%H:%M:%S")
    log_text.insert(tk.END, f"[{timestamp}] {msg}\n")
    log_text.see(tk.END)  # Auto-scroll to bottom
    root.update()


def show_success(result):
    btn_run.config(state="normal", text="Run Process")
    progress_label.config(text="✓ Complete!", fg="green")

    # Add completion message to log
    log_text.insert(tk.END, f"\n{'=' * 80}\n")
    log_text.insert(tk.END, f"✓ PROCESSING COMPLETE!\n")
    log_text.insert(tk.END, f"Exported {len(result['exported_files'])} files.\n")
    if result["no_match_inputs"]:
        log_text.insert(tk.END, f"No matches for dates: {', '.join(result['no_match_inputs'])}\n")
    log_text.insert(tk.END, f"{'=' * 80}\n")
    log_text.see(tk.END)

    msg = f"✓ Processing complete!\n\nExported {len(result['exported_files'])} files."
    if result["no_match_inputs"]:
        msg += "\n\nNo matching records found for the following dates:\n" + ", ".join(result["no_match_inputs"])
    messagebox.showinfo("Complete", msg)


def show_error(error_msg):
    btn_run.config(state="normal", text="Run Process")
    progress_label.config(text="✗ Error occurred", fg="red")

    # Add error to log
    log_text.insert(tk.END, f"\n{'=' * 80}\n")
    log_text.insert(tk.END, f"✗ ERROR: {error_msg}\n")
    log_text.insert(tk.END, f"{'=' * 80}\n")
    log_text.see(tk.END)

    messagebox.showerror("Error", error_msg)


# Build GUI
root = tk.Tk()
root.title("Thermo Shipment Automation Tool")
root.geometry("950x600")  # 增加高度以容纳日志窗口

# Set style
style = ttk.Style()
style.theme_use('clam')

# Title
title_label = tk.Label(root, text="Thermo Shipment Automation Tool", font=("Arial", 14, "bold"), fg="#2c3e50")
title_label.grid(row=0, column=0, columnspan=3, pady=10)

# Database file
tk.Label(root, text="Database File (Thermo Database V10.xlsx):").grid(row=1, column=0, sticky="e", padx=6, pady=6)
entry_db = tk.Entry(root, width=80)
entry_db.grid(row=1, column=1, padx=6, pady=6)
tk.Button(root, text="Browse", command=lambda: entry_db.delete(0, tk.END) or entry_db.insert(0,
                                                                                             filedialog.askopenfilename(
                                                                                                 title="Select Thermo Database V10.xlsx",
                                                                                                 filetypes=[
                                                                                                     ("Excel files",
                                                                                                      "*.xlsx;*.xls;*.xlsm")]))).grid(
    row=1, column=2, padx=6, pady=6)

# Shipment file
tk.Label(root, text="Shipment File (Thermo Dec shipment.xlsx):").grid(row=2, column=0, sticky="e", padx=6, pady=6)
entry_ship = tk.Entry(root, width=80)
entry_ship.grid(row=2, column=1, padx=6, pady=6)
tk.Button(root, text="Browse", command=lambda: entry_ship.delete(0, tk.END) or entry_ship.insert(0,
                                                                                                 filedialog.askopenfilename(
                                                                                                     title="Select Thermo Dec shipment.xlsx",
                                                                                                     filetypes=[
                                                                                                         ("Excel files",
                                                                                                          "*.xlsx;*.xls;*.xlsm")]))).grid(
    row=2, column=2, padx=6, pady=6)

# Date input
tk.Label(root, text="Enter Date(s) (comma separated):").grid(row=3, column=0, sticky="e", padx=6, pady=6)
entry_dates = tk.Entry(root, width=80)
entry_dates.grid(row=3, column=1, padx=6, pady=6)
tk.Label(root, text="Example: 2025.12.12 (yyyy.mm.dd)", fg="gray").grid(row=3, column=2, padx=6, pady=6)

# Output folder
tk.Label(root, text="Output Folder:").grid(row=4, column=0, sticky="e", padx=6, pady=6)
entry_out = tk.Entry(root, width=80)
entry_out.grid(row=4, column=1, padx=6, pady=6)
tk.Button(root, text="Browse", command=lambda: entry_out.delete(0, tk.END) or entry_out.insert(0,
                                                                                               filedialog.askdirectory(
                                                                                                   title="Select output folder"))).grid(
    row=4, column=2, padx=6, pady=6)

# Progress label
progress_label = tk.Label(root, text="", fg="blue", font=("Arial", 9))
progress_label.grid(row=5, column=1, pady=6)

# Run button
btn_run = tk.Button(root, text="Run Process", bg="#4CAF50", fg="white", width=20, height=2, font=("Arial", 10, "bold"),
                    command=on_run)
btn_run.grid(row=6, column=1, pady=10)

# Log window section
tk.Label(root, text="Process Log:", font=("Arial", 10, "bold")).grid(row=7, column=0, columnspan=3, sticky="w", padx=6,
                                                                     pady=(10, 0))

# Create frame for log text widget with scrollbar
log_frame = tk.Frame(root)
log_frame.grid(row=8, column=0, columnspan=3, padx=6, pady=6, sticky="nsew")

# Configure grid weights for resizing
root.grid_rowconfigure(8, weight=1)
root.grid_columnconfigure(1, weight=1)

# Scrollbar
log_scrollbar = tk.Scrollbar(log_frame)
log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Text widget for logs
log_text = tk.Text(log_frame, height=12, width=110, yscrollcommand=log_scrollbar.set,
                   font=("Consolas", 9), bg="#f5f5f5", fg="#333333")
log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
log_scrollbar.config(command=log_text.yview)

# Instructions
info_text = "Note: Processing may take several minutes depending on data size. Check the log window for detailed progress."
tk.Label(root, text=info_text, justify="left", fg="gray", font=("Arial", 8)).grid(row=9, column=0, columnspan=3, pady=6)

root.mainloop()